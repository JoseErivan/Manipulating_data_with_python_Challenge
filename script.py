import openpyxl
import math
import sys

""" 
    Desenvolvedor: José Erivan Ramalho Ferreira Filho

    Programa para manipulação de dados em uma tabela excell, 
    utiliza os valores das células para calcular os valores que faltam
    e salvá-lo na tabela.
"""

"""
    Developer: José Erivan Ramalho Ferreira Filho

    Program for manipulating data in an Excel table,
    uses cell values to calculate missing values
    and save it to the table.
"""


def find_situation(p1, p2, p3, fouls):
    # Find the average.
    average = ((p1 + p2 + p3) / 3) / 10

    if fouls > 15:
        return "Reprovado por falta", 0

    # Check the average value.
    elif average < 5:
        return "Reprovado por nota", 0

    elif average >= 5 and average < 7:
        naf = (average / 2) * 10  # Calculating naf.
        return "Exame final", math.ceil(naf)

    else:
        return "Aprovado", 0


# Finds the student's situation and puts the results in the table.
def situation(rows):
    # Passing through all lines.
    for row in rows:
        name = row[1]
        fouls = row[2]
        p1 = row[3]
        p2 = row[4]
        p3 = row[5]

        final_situation, naf = find_situation(
            p1, p2, p3, fouls)  # Finding the values

        # Placing the values in their respective locations.
        spreadsheet.cell(row=row[0] + 3, column=7,
                         value=final_situation)
        spreadsheet.cell(row=row[0] + 3, column=8,
                         value=naf)

    workbook.save(spreadsheetName)


# Checking if the script was used correctly.
if len(sys.argv) - 1 != 1:
    print("Formato de uso: python ./script.py spreadsheetName")
    sys.exit(1)

# Saving the spreadsheet on a variable.
spreadsheetName = sys.argv[1]

# Opening the spreadsheet.
print("Abrindo a tabela de dados...")
workbook = openpyxl.load_workbook(spreadsheetName)

# Separating the right spreadsheet.
spreadsheet = workbook.worksheets[0]

try:
    # Taking the rows values.
    rows = list(spreadsheet.iter_rows(values_only=True, min_row=4))

    print("Fazendo cálculos necessários e adicionando na tabela...")

    situation(rows)
    print("Finalizando o programa...")

except FileNotFoundError:
    # Handling if the file is not found.
    print(f"A tabela '{spreadsheet}' não foi encontrada.")
except Exception as e:
    # Handling if an error occurs.
    print(f"Um erro aconteceu na abertura do arquivo: {e}")
